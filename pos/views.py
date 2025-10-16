# pos/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from .models import Product, Sale, SaleItem, Customer, SaleReturn,SaleReturnItem
from django.contrib.auth import logout
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Q, Count, Avg
from .models import CashDrawerSession
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db import transaction
from datetime import datetime
import io
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors


@login_required
def home_dispatch_view(request):
    """Redirige a cada usuario según su rol - VERSIÓN CORREGIDA"""
    if request.user.is_staff or request.user.is_superuser:
        return redirect('admin_dashboard')
    return redirect('pos_main')


@login_required
def pos_view(request):
    """Vista principal del POS - ahora con carrito Y sesión activa"""
    # Obtener sesión activa del usuario
    active_session = get_active_session(request.user)

    # Cargar items del carrito si existen
    cart_items = []
    total = 0

    if 'cart' in request.session:
        cart = request.session['cart']
        for item in cart:
            try:
                product = Product.objects.get(id=item['product_id'])
                cart_items.append({
                    'name': item['name'],
                    'sku': item['sku'],
                    'price': item['price'],
                    'quantity': item['quantity'],
                    'subtotal': float(item['price']) * item['quantity'],
                    'product_obj': product  # Para acceder al stock en template
                })
                total += float(item['price']) * item['quantity']
            except Product.DoesNotExist:
                # Si el producto fue eliminado, saltarlo
                continue

    return render(request, 'pos/pos_main.html', {
        'cart_items': cart_items,
        'total': total,
        'active_session': active_session
    })


@login_required
@csrf_exempt
def add_product_view(request):
    """Vista HTMX para añadir productos al carrito - CON VALIDACIÓN DE STOCK"""
    try:
        if request.method == "POST":
            sku = request.POST.get('sku', '').strip()

            if not sku:
                return HttpResponse('<tr><td colspan="4" class="error-message">Por favor ingresa un SKU</td></tr>')

            try:
                product = Product.objects.get(sku=sku)

                # ✅ NUEVO: VALIDAR STOCK (Sprint 2)
                if product.stock <= 0:
                    return HttpResponse(
                        '<tr><td colspan="4" class="error-message">❌ Producto sin stock disponible</td></tr>'
                    )

                # Inicializar carrito en sesión si no existe
                if 'cart' not in request.session:
                    request.session['cart'] = []

                # Buscar si el producto ya está en el carrito
                cart = request.session['cart']
                product_found = False

                for item in cart:
                    if item['product_id'] == product.id:
                        # ✅ NUEVO: Validar que no exceda el stock disponible
                        if item['quantity'] + 1 > product.stock:
                            return HttpResponse(
                                f'<tr><td colspan="4" class="error-message">❌ No hay suficiente stock. Stock disponible: {product.stock}</td></tr>'
                            )
                        item['quantity'] += 1
                        product_found = True
                        break

                # Si no está, añadirlo
                if not product_found:
                    cart.append({
                        'product_id': product.id,
                        'name': product.name,
                        'sku': product.sku,
                        'price': str(product.price),
                        'quantity': 1
                    })

                # Guardar carrito en sesión
                request.session.modified = True

                # ✅ NUEVO: Renderizar fila del producto CON INFO DE STOCK
                subtotal = float(product.price) * 1
                stock_class = "no-stock" if product.stock == 0 else "low-stock" if product.stock < 10 else ""
                stock_text = f"SIN STOCK" if product.stock == 0 else f"Stock bajo: {product.stock} unidades" if product.stock < 10 else f"Stock disponible: {product.stock} unidades"

                html_response = f"""
                <tr class="success-row">
                    <td>
                        {product.name} (SKU: {product.sku})
                        <div class="stock-info">
                            <span class="{stock_class}">{stock_text}</span>
                        </div>
                    </td>
                    <td>${product.price}</td>
                    <td>1</td>
                    <td>${subtotal:.2f}</td>
                </tr>
                """
                return HttpResponse(html_response)

            except Product.DoesNotExist:
                return HttpResponse('<tr><td colspan="4" class="error-message">❌ Producto no encontrado</td></tr>')

        return HttpResponse('Método no permitido', status=405)

    except Exception as e:
        return HttpResponse(f'<tr><td colspan="4" class="error-message">Error: {str(e)}</td></tr>')


@login_required
@transaction.atomic
def checkout_view(request):
    """Vista para finalizar la venta - AHORA CON CLIENTES"""
    try:
        if request.method == "POST":
            payment_method = request.POST.get('payment_method', 'cash')
            customer_id = request.POST.get('customer_id')  # ✅ NUEVO: Obtener cliente
            active_session = get_active_session(request.user)

            if not active_session:
                messages.error(request, "❌ No tienes una sesión de caja activa")
                return redirect('open_session')

            cart = request.session.get('cart', [])
            if not cart:
                messages.error(request, "El carrito está vacío")
                return redirect('pos_main')

            # Obtener el cliente si se seleccionó uno
            customer = None
            if customer_id and customer_id != '':
                try:
                    customer = Customer.objects.get(id=customer_id)
                except Customer.DoesNotExist:
                    messages.warning(request, "Cliente no encontrado, continuando sin cliente")

            # Dentro de la transacción atómica
            with transaction.atomic():
                # Validar stock
                for item in cart:
                    product = Product.objects.select_for_update().get(id=item['product_id'])
                    if product.stock < item['quantity']:
                        messages.error(
                            request,
                            f"Stock insuficiente para {product.name}. Disponible: {product.stock}, Solicitado: {item['quantity']}"
                        )
                        return redirect('pos_main')

                total_amount = sum(float(item['price']) * item['quantity'] for item in cart)

                # Crear venta CON CLIENTE
                sale = Sale.objects.create(
                    total_amount=total_amount,
                    cash_drawer_session=active_session,
                    payment_method=payment_method,
                    customer=customer  # ✅ NUEVO: Asignar cliente
                )

                # Reducir stock y crear items
                for item in cart:
                    product = Product.objects.get(id=item['product_id'])
                    product.stock -= item['quantity']
                    product.save()

                    SaleItem.objects.create(
                        sale=sale,
                        product=product,
                        product_name=product.name,
                        quantity=item['quantity'],
                        unit_price=item['price']
                    )

            # Limpiar carrito (fuera de la transacción)
            request.session['cart'] = []
            request.session.modified = True

            # Mensaje de confirmación
            if customer:
                messages.success(request,
                                 f"✅ Venta #{sale.id} registrada para {customer.name} - Total: ${total_amount:.2f}")
            else:
                messages.success(request, f"✅ Venta #{sale.id} registrada - Total: ${total_amount:.2f}")

            return redirect('pos_main')

    except Exception as e:
        messages.error(request, f"Error al procesar la venta: {str(e)}")
        return redirect('pos_main')

    return redirect('pos_main')


@login_required
def search_customers_view(request):
    """Vista HTMX para buscar clientes - VERSIÓN CORREGIDA"""
    # Obtener query de diferentes posibles nombres de parámetro
    query = request.GET.get('q', '').strip()
    if not query:
        query = request.GET.get('customer_search', '').strip()

    print(f"🔍 Búsqueda de clientes: '{query}'")  # Para debug

    if query:
        # Buscar por nombre, RUC o email
        customers = Customer.objects.filter(
            Q(name__icontains=query) |
            Q(tax_id__icontains=query) |
            Q(email__icontains=query)
        )[:10]  # Limitar a 10 resultados

        print(f"📋 Clientes encontrados: {customers.count()}")  # Para debug
    else:
        customers = Customer.objects.none()

    return render(request, 'pos/partials/customer_search_results.html', {
        'customers': customers,
        'query': query
    })


@require_http_methods(["GET", "POST"])
def custom_logout_view(request):
    """Vista personalizada para logout que acepta GET y POST"""
    logout(request)
    return redirect('login')


@login_required
def open_session_view(request):
    """Vista para abrir una nueva sesión de caja"""
    # Verificar si ya tiene una sesión activa
    active_session = CashDrawerSession.objects.filter(
        user=request.user,
        end_time__isnull=True
    ).first()

    if active_session:
        return redirect('pos_main')

    if request.method == 'POST':
        starting_balance = request.POST.get('starting_balance')
        try:
            starting_balance = float(starting_balance)
            if starting_balance >= 0:
                # Crear nueva sesión
                session = CashDrawerSession.objects.create(
                    user=request.user,
                    starting_balance=starting_balance
                )
                messages.success(request, f"✅ Caja abierta con fondo inicial: ${starting_balance:.2f}")
                return redirect('pos_main')
            else:
                messages.error(request, "El fondo inicial no puede ser negativo")
        except (ValueError, TypeError):
            messages.error(request, "Por favor ingresa un monto válido")

    return render(request, 'pos/open_session.html')


@login_required
def close_session_view(request):
    """Vista para cerrar la sesión de caja activa"""
    try:
        # Obtener sesión activa
        active_session = get_active_session(request.user)

        if not active_session:
            messages.error(request, "No tienes una sesión de caja activa")
            return redirect('pos_main')

        # Calcular ventas
        cash_sales_result = active_session.sales.filter(payment_method='cash').aggregate(
            total=Sum('total_amount')
        )
        cash_sales = cash_sales_result['total'] or 0

        card_sales_result = active_session.sales.filter(payment_method='card').aggregate(
            total=Sum('total_amount')
        )
        card_sales = card_sales_result['total'] or 0

        total_sales_result = active_session.sales.aggregate(
            total=Sum('total_amount')
        )
        total_sales = total_sales_result['total'] or 0

        expected_cash = active_session.starting_balance + cash_sales

        if request.method == 'POST':
            ending_balance = request.POST.get('ending_balance')
            notes = request.POST.get('notes', '')

            try:
                ending_balance = float(ending_balance)
                if ending_balance >= 0:
                    # Cerrar sesión
                    active_session.end_time = timezone.now()
                    active_session.ending_balance = ending_balance
                    active_session.notes = notes
                    active_session.save()

                    # Calcular diferencia para el mensaje
                    difference = ending_balance - expected_cash
                    if difference == 0:
                        msg = "✅ Caja cerrada perfectamente - Sin diferencias"
                    elif difference > 0:
                        msg = f"✅ Caja cerrada - Sobrante: ${difference:.2f}"
                    else:
                        msg = f"⚠️ Caja cerrada - Faltante: ${abs(difference):.2f}"

                    messages.success(request, msg)
                    return redirect('logout')
                else:
                    messages.error(request, "El saldo final no puede ser negativo")
            except (ValueError, TypeError):
                messages.error(request, "Por favor ingresa un monto válido")

        context = {
            'active_session': active_session,
            'cash_sales': cash_sales,
            'card_sales': card_sales,
            'total_sales': total_sales,
            'expected_cash': expected_cash,
        }

        return render(request, 'pos/close_session.html', context)

    except Exception as e:
        messages.error(request, f"Error al procesar el cierre: {str(e)}")
        return redirect('pos_main')


def get_active_session(user):
    """Función auxiliar para obtener sesión activa"""
    try:
        return CashDrawerSession.objects.filter(
            user=user,
            end_time__isnull=True
        ).first()
    except CashDrawerSession.DoesNotExist:
        return None


@staff_member_required
def admin_dashboard(request):
    """Dashboard del Administrador - MEJORADO"""
    today = timezone.now().date()

    # Métricas principales
    today_sales = Sale.objects.filter(created_at__date=today).aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    today_transactions = Sale.objects.filter(created_at__date=today).count()
    today_avg_ticket = today_sales / today_transactions if today_transactions > 0 else 0
    active_sessions = CashDrawerSession.objects.filter(end_time__isnull=True).count()

    # Métricas de la semana
    week_ago = today - timezone.timedelta(days=7)
    week_sales = Sale.objects.filter(created_at__date__gte=week_ago).aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    # Ventas por método de pago hoy
    cash_sales_today = Sale.objects.filter(
        created_at__date=today,
        payment_method='cash'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    card_sales_today = Sale.objects.filter(
        created_at__date=today,
        payment_method='card'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Productos más vendidos
    top_products = SaleItem.objects.values(
        'product_name'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('unit_price')
    ).order_by('-total_sold')[:5]

    # Sesiones de hoy
    today_sessions = CashDrawerSession.objects.filter(
        start_time__date=today
    ).select_related('user')

    # Ventas recientes
    recent_sales = Sale.objects.select_related(
        'cash_drawer_session',
        'cash_drawer_session__user'
    ).order_by('-created_at')[:10]

    # Productos con stock bajo
    low_stock_products = Product.objects.filter(stock__lt=10, stock__gt=0).order_by('stock')[:5]
    out_of_stock_products = Product.objects.filter(stock=0)[:5]

    context = {
        # Métricas principales
        'today_sales': today_sales,
        'today_transactions': today_transactions,
        'today_avg_ticket': today_avg_ticket,
        'active_sessions': active_sessions,

        # Nuevas métricas
        'week_sales': week_sales,
        'cash_sales_today': cash_sales_today,
        'card_sales_today': card_sales_today,

        # Datos detallados
        'top_products': top_products,
        'today_sessions': today_sessions,
        'recent_sales': recent_sales,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'today_date': today,
        'week_ago': week_ago,
    }

    return render(request, 'pos/admin_dashboard.html', context)


@staff_member_required
def sales_report_view(request):
    """Reporte de ventas por rango de fechas - CON EXPORTACIÓN"""
    sales = Sale.objects.none()
    total_sales = 0
    total_transactions = 0
    start_date = None
    end_date = None

    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        # VERIFICAR SI ES UNA EXPORTACIÓN
        if 'export_excel' in request.POST or 'export_pdf' in request.POST:
            if start_date_str and end_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

                    sales = Sale.objects.filter(
                        created_at__date__range=[start_date, end_date]
                    ).select_related(
                        'cash_drawer_session',
                        'cash_drawer_session__user'
                    ).order_by('-created_at')

                    if 'export_excel' in request.POST:
                        return generate_excel_report(sales, start_date, end_date)
                    elif 'export_pdf' in request.POST:
                        return generate_pdf_report(sales, start_date, end_date)

                except ValueError:
                    messages.error(request, "Formato de fecha inválido")

        # PROCESAMIENTO NORMAL DEL REPORTE
        elif start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

                sales = Sale.objects.filter(
                    created_at__date__range=[start_date, end_date]
                ).select_related(
                    'cash_drawer_session',
                    'cash_drawer_session__user'
                ).order_by('-created_at')

                if sales.exists():
                    total_sales_result = sales.aggregate(total=Sum('total_amount'))
                    total_sales = total_sales_result['total'] or 0
                    total_transactions = sales.count()

            except ValueError:
                messages.error(request, "Formato de fecha inválido")

    context = {
        'sales': sales,
        'total_sales': total_sales,
        'total_transactions': total_transactions,
        'start_date': start_date,
        'end_date': end_date,
        'today': timezone.now().date(),
    }

    return render(request, 'pos/sales_report.html', context)


def generate_excel_report(sales, start_date, end_date):
    """Generar reporte en formato Excel"""
    # Crear libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte Ventas {start_date} a {end_date}"

    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Ventas - {start_date} a {end_date}"
    ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)

    # Encabezados de tabla
    headers = ['ID Venta', 'Fecha', 'Vendedor', 'Método Pago', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Datos
    for row, sale in enumerate(sales, 4):
        ws.cell(row=row, column=1, value=sale.id)
        ws.cell(row=row, column=2, value=sale.created_at.strftime('%d/%m/%Y %H:%M'))

        # Vendedor
        vendedor = "N/A"
        if sale.cash_drawer_session and sale.cash_drawer_session.user:
            vendedor = sale.cash_drawer_session.user.username
        ws.cell(row=row, column=3, value=vendedor)

        # Método de pago
        metodo_pago = 'Efectivo' if sale.payment_method == 'cash' else 'Tarjeta'
        ws.cell(row=row, column=4, value=metodo_pago)

        ws.cell(row=row, column=5, value=float(sale.total_amount))

    # Ajustar anchos de columna
    column_widths = [12, 20, 20, 15, 12]
    for col, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

    # Totales
    total_row = len(sales) + 5
    ws.cell(row=total_row, column=4, value="TOTAL:").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row - 1})").font = openpyxl.styles.Font(bold=True)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_ventas_{start_date}_a_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def generate_pdf_report(sales, start_date, end_date):
    """Generar reporte en formato PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()

    # Título
    title = Paragraph(f"<b>Reporte de Ventas</b><br/>Período: {start_date} a {end_date}", styles['Heading2'])
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    # Datos de resumen
    total_sales = sum(sale.total_amount for sale in sales)
    total_transactions = len(sales)

    summary_data = [
        ["Total de Ventas:", f"${total_sales:.2f}"],
        ["Total de Transacciones:", str(total_transactions)],
        ["Período:", f"{start_date} a {end_date}"]
    ]

    summary_table = Table(summary_data, colWidths=[150, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5C5CBD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e9ecef')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))

    # Tabla de ventas detalladas
    if sales:
        data = [['ID', 'Fecha', 'Vendedor', 'Método', 'Total']]

        for sale in sales:
            vendedor = "N/A"
            if sale.cash_drawer_session and sale.cash_drawer_session.user:
                vendedor = sale.cash_drawer_session.user.username

            metodo = 'Efectivo' if sale.payment_method == 'cash' else 'Tarjeta'

            data.append([
                str(sale.id),
                sale.created_at.strftime('%d/%m/%Y %H:%M'),
                vendedor,
                metodo,
                f"${sale.total_amount:.2f}"
            ])

        # Añadir fila de total
        data.append(['', '', '', 'TOTAL:', f"${total_sales:.2f}"])

        table = Table(data, colWidths=[50, 80, 80, 60, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5C5CBD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

    # Construir PDF
    doc.build(elements)

    # Preparar respuesta
    response = HttpResponse(content_type='application/pdf')
    filename = f"reporte_ventas_{start_date}_a_{end_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


@login_required
def returns_main_view(request):
    """Vista principal de devoluciones"""
    return render(request, 'pos/returns_main.html')


@login_required
def search_sale_for_return_view(request):
    """Buscar venta para devolución"""
    sale_id = request.GET.get('sale_id', '').strip()
    sale = None
    error = None

    if sale_id:
        try:
            sale = Sale.objects.select_related(
                'customer',
                'cash_drawer_session',
                'cash_drawer_session__user'
            ).prefetch_related('items').get(id=sale_id)
        except Sale.DoesNotExist:
            error = f"❌ No se encontró la venta #{sale_id}"
        except ValueError:
            error = "❌ ID de venta inválido"

    return render(request, 'pos/partials/sale_return_details.html', {
        'sale': sale,
        'error': error,
        'sale_id': sale_id
    })


@login_required
@transaction.atomic
def process_return_view(request):
    """Procesar la devolución"""
    if request.method == 'POST':
        sale_id = request.POST.get('sale_id')
        reason = request.POST.get('reason', '')

        try:
            sale = Sale.objects.get(id=sale_id)

            # Crear la devolución
            sale_return = SaleReturn.objects.create(
                original_sale=sale,
                reason=reason,
                processed_by=request.user
            )

            total_refund = 0
            return_items = []

            # Procesar cada item devuelto
            for key, value in request.POST.items():
                if key.startswith('return_qty_'):
                    item_id = key.replace('return_qty_', '')
                    return_qty = int(value)

                    if return_qty > 0:
                        try:
                            sale_item = SaleItem.objects.get(id=item_id, sale=sale)

                            # Validar que no se devuelva más de lo comprado
                            if return_qty <= sale_item.quantity:
                                # Crear item de devolución
                                return_item = SaleReturnItem(
                                    return_request=sale_return,
                                    product=sale_item.product,
                                    quantity=return_qty,
                                    unit_price=sale_item.unit_price
                                )
                                return_items.append(return_item)

                                # Revertir stock
                                product = sale_item.product
                                product.stock += return_qty
                                product.save()

                                # Calcular reembolso
                                total_refund += return_qty * sale_item.unit_price
                            else:
                                messages.error(request,
                                               f"No se puede devolver más de {sale_item.quantity} unidades de {sale_item.product_name}")
                                return redirect('returns_main')

                        except SaleItem.DoesNotExist:
                            messages.error(request, "Item de venta no encontrado")
                            return redirect('returns_main')

            # Guardar todos los items de devolución
            if return_items:
                SaleReturnItem.objects.bulk_create(return_items)
                sale_return.total_refund = total_refund
                sale_return.save()

                messages.success(request, f"✅ Devolución procesada exitosamente. Reembolso: ${total_refund:.2f}")
                return redirect('returns_main')
            else:
                messages.error(request, "❌ No se seleccionaron productos para devolver")
                return redirect('returns_main')

        except Sale.DoesNotExist:
            messages.error(request, "Venta no encontrada")
            return redirect('returns_main')
        except Exception as e:
            messages.error(request, f"Error al procesar devolución: {str(e)}")
            return redirect('returns_main')

    return redirect('returns_main')